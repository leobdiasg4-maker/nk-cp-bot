import os
import json
import logging
from datetime import datetime, date
from zoneinfo import ZoneInfo
from dateutil.relativedelta import relativedelta

import gspread
from google.oauth2.service_account import Credentials
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters
)
import openpyxl

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(format="%(asctime)s [%(levelname)s] %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)

# ── Constantes ────────────────────────────────────────────────────────────────
BOT_TOKEN      = os.environ["BOT_TOKEN"]
SHEET_ID       = os.environ["SHEET_ID"]
MY_TELEGRAM_ID = int(os.environ.get("MY_TELEGRAM_ID", "647725027"))
TZ             = ZoneInfo("America/Sao_Paulo")

_DEFAULT = {
    "Empresa":   ["NK Soluções", "NK Pré-Moldados"],
    "Categoria": ["DAS","INSS","FGTS","Folha","Fornecedor","Aluguel","Empréstimo","Honorários","Outros"],
    "Conta":     ["Nubank PJ","Sicoob","C6","PagBank","Santander NK Soluções","Santander NK Pré-Moldados","Cora","Dinheiro"],
}
FREQ_OPT = ["Única", "Mensal", "Semanal", "Anual"]

COLS = [
    "ID", "Empresa", "Categoria", "Descrição", "Credor",
    "Valor (R$)", "Vencimento", "Status", "Data Pagamento",
    "Valor Pago", "Conta Bancária", "Recorrente", "Frequência",
    "Observação", "Lançado por", "Criado em", "Cancelado", "Motivo Cancelamento"
]

(S_EMPRESA, S_CATEGORIA, S_DESCRICAO, S_CREDOR, S_VALOR,
 S_VENCIMENTO, S_RECORRENTE, S_FREQUENCIA, S_OBS, S_CONFIRMA) = range(10)

# ── Google Sheets ─────────────────────────────────────────────────────────────
def get_sheet():
    raw = os.environ["GOOGLE_CREDENTIALS"]
    if not raw.strip().startswith("{"):
        raw = "{" + raw + "}"
    info = json.loads(raw)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)

def get_config_list(tipo: str) -> list:
    try:
        sh  = get_sheet()
        w   = sh.worksheet("Config")
        rows = w.get_all_values()[1:]
        vals = [r[1] for r in rows if len(r) >= 2 and r[0] == tipo and r[1]]
        return vals if vals else _DEFAULT.get(tipo, [])
    except Exception:
        return _DEFAULT.get(tipo, [])

def ensure_sheets(sh):
    existing = [w.title for w in sh.worksheets()]
    if "Contas" not in existing:
        w = sh.add_worksheet("Contas", rows=1000, cols=len(COLS))
        w.append_row(COLS)
    if "Config" not in existing:
        w = sh.add_worksheet("Config", rows=200, cols=2)
        rows = [["Tipo", "Valor"]]
        for tipo, vals in _DEFAULT.items():
            for v in vals:
                rows.append([tipo, v])
        w.update("A1", rows)
    else:
        # Migra formato antigo se necessário
        w = sh.worksheet("Config")
        header = w.row_values(1)
        if header != ["Tipo", "Valor"]:
            rows = [["Tipo", "Valor"]]
            for tipo, vals in _DEFAULT.items():
                for v in vals:
                    rows.append([tipo, v])
            w.clear()
            w.update("A1", rows)
    if "Log" not in existing:
        w = sh.add_worksheet("Log", rows=1000, cols=4)
        w.append_row(["Timestamp", "Ação", "ID Conta", "Usuário"])
    if "Usuarios" not in existing:
        w = sh.add_worksheet("Usuarios", rows=50, cols=3)
        w.append_row(["login", "senha", "nome"])
        w.append_row(["leonardo", "NK2026", "Leonardo"])
        w.append_row(["nicanor", "NK2026", "Nicanor"])

def next_id(ws) -> str:
    vals = [int(v.replace("CP","")) for v in ws.col_values(1)[1:] if v.startswith("CP")]
    return f"CP{(max(vals)+1 if vals else 1):04d}"

def get_all_contas(ws):
    rows = ws.get_all_values()
    if len(rows) <= 1:
        return []
    headers = rows[0]
    result = []
    for r in rows[1:]:
        d = dict(zip(headers, r))
        if d.get("Cancelado") == "Sim":
            continue  # ignora canceladas nas listagens normais
        result.append(d)
    return result

def get_all_contas_incluindo_canceladas(ws):
    rows = ws.get_all_values()
    if len(rows) <= 1:
        return []
    headers = rows[0]
    return [dict(zip(headers, r)) for r in rows[1:]]

def append_conta(ws, data: dict):
    ws.append_row([data.get(c, "") for c in COLS])

def update_status(ws, cp_id, new_status, data_pag="", valor_pago=""):
    rows = ws.get_all_values()
    for i, row in enumerate(rows[1:], start=2):
        if row[0] == cp_id:
            ws.update_cell(i, 8, new_status)
            if data_pag:  ws.update_cell(i, 9, data_pag)
            if valor_pago: ws.update_cell(i, 10, valor_pago)
            return i, row
    return None, None

# ── Helpers ───────────────────────────────────────────────────────────────────
def fmt_brl(val):
    try:
        return f"R$ {float(str(val).replace(',','.') or 0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except Exception:
        return str(val)

def parse_date(s: str):
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None

def days_until(venc_str: str):
    d = parse_date(venc_str)
    if not d:
        return None
    return (d - date.today()).days

def next_vencimento(venc_str: str, freq: str) -> str:
    d = parse_date(venc_str)
    if not d:
        return ""
    if freq == "Mensal":   d = d + relativedelta(months=1)
    elif freq == "Semanal": d = d + relativedelta(weeks=1)
    elif freq == "Anual":   d = d + relativedelta(years=1)
    return d.strftime("%d/%m/%Y")

def status_emoji(status, days):
    if status == "Pago": return "✅"
    if days is None: return "❓"
    if days < 0:  return "🔴"
    if days == 0: return "🚨"
    if days <= 3: return "🟠"
    if days <= 7: return "🟡"
    return "🟢"

def kb(options, columns=2):
    rows = [options[i:i+columns] for i in range(0, len(options), columns)]
    return ReplyKeyboardMarkup(rows, one_time_keyboard=True, resize_keyboard=True)

def is_authorized(update: Update):
    return update.effective_user.id == MY_TELEGRAM_ID

# ── Resumo ─────────────────────────────────────────────────────────────────────
def build_resumo(contas, titulo: str) -> str:
    if not contas:
        return f"📋 *{titulo}*\n\nNenhuma conta encontrada."
    total_pend = sum(
        float(str(c.get("Valor (R$)","0")).replace(",",".") or 0)
        for c in contas if c.get("Status") in ("Pendente","Atrasado","Parcial")
    )
    atrasadas  = [c for c in contas if days_until(c.get("Vencimento","")) is not None and days_until(c.get("Vencimento","")) < 0 and c.get("Status") != "Pago"]
    vence_hoje = [c for c in contas if days_until(c.get("Vencimento","")) == 0 and c.get("Status") != "Pago"]
    vence_7    = [c for c in contas if days_until(c.get("Vencimento","")) is not None and 1 <= days_until(c.get("Vencimento","")) <= 7 and c.get("Status") != "Pago"]

    lines = [f"📋 *{titulo}*\n"]
    lines.append(f"💰 Total pendente: *{fmt_brl(total_pend)}*")
    lines.append(f"🔴 Atrasadas: {len(atrasadas)} | 🚨 Hoje: {len(vence_hoje)} | 🟡 7 dias: {len(vence_7)}\n")

    def bloco(label, lista):
        if not lista: return
        lines.append(f"*{label}*")
        for c in lista:
            d  = days_until(c.get("Vencimento",""))
            em = status_emoji(c.get("Status",""), d)
            lines.append(f"{em} {c.get('ID','')} | {c.get('Empresa','')} | {c.get('Credor','')} | {fmt_brl(c.get('Valor (R$)',0))} | {c.get('Vencimento','')}")
        lines.append("")

    bloco("🔴 Atrasadas", atrasadas)
    bloco("🚨 Vencem hoje", vence_hoje)
    bloco("🟡 Próximos 7 dias", vence_7)
    return "\n".join(lines)

# ── Jobs ───────────────────────────────────────────────────────────────────────
async def job_alertas(context: ContextTypes.DEFAULT_TYPE):
    try:
        sh = get_sheet()
        ws = sh.worksheet("Contas")
        contas = get_all_contas(ws)
        alertas = [(days_until(c.get("Vencimento","")), c) for c in contas
                   if c.get("Status") != "Pago" and days_until(c.get("Vencimento","")) in (0,1,3,7)]
        if not alertas: return
        msg = "⏰ *Alertas de Vencimento*\n\n"
        for d, c in sorted(alertas, key=lambda x: x[0] or 0):
            em    = status_emoji(c.get("Status",""), d)
            prazo = "HOJE" if d == 0 else f"em {d} dia(s)"
            msg  += f"{em} *{c.get('ID','')}* — {c.get('Credor','')} ({c.get('Empresa','')})\n   {fmt_brl(c.get('Valor (R$)',0))} — vence *{prazo}* ({c.get('Vencimento','')})\n\n"
        await context.bot.send_message(chat_id=MY_TELEGRAM_ID, text=msg, parse_mode="Markdown")
    except Exception as e:
        log.error("job_alertas: %s", e)

async def job_resumo_diario(context: ContextTypes.DEFAULT_TYPE):
    try:
        sh = get_sheet()
        ws = sh.worksheet("Contas")
        contas = [c for c in get_all_contas(ws) if c.get("Status") != "Pago"]
        msg = build_resumo(contas, f"Resumo Diário — {date.today().strftime('%d/%m/%Y')}")
        await context.bot.send_message(chat_id=MY_TELEGRAM_ID, text=msg, parse_mode="Markdown")
    except Exception as e:
        log.error("job_resumo_diario: %s", e)

# ── /start ─────────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    await update.message.reply_text(
        "👋 *NK Contas a Pagar*\n\n"
        "*/nova* — Lançar conta (guiado)\n"
        "*/cp* — Lançar conta (rápido)\n"
        "*/pagar* — Marcar como pago\n"
        "*/listar* — Contas pendentes\n"
        "*/resumo* — Resumo geral\n"
        "*/resumo\\_dia* | */resumo\\_semana* | */resumo\\_mes*\n"
        "*/ajuda* — Todos os comandos\n\n"
        "📎 Envie um `.xlsx` para importar em lote.",
        parse_mode="Markdown"
    )

# ── /nova (guiado) ─────────────────────────────────────────────────────────────
async def cmd_nova(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return ConversationHandler.END
    context.user_data.clear()
    empresas = get_config_list("Empresa")
    await update.message.reply_text("📝 *Nova conta a pagar*\n\nQual empresa?",
        reply_markup=kb(empresas), parse_mode="Markdown")
    return S_EMPRESA

async def step_empresa(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Empresa"] = update.message.text
    cats = get_config_list("Categoria")
    await update.message.reply_text("Categoria:", reply_markup=kb(cats))
    return S_CATEGORIA

async def step_categoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Categoria"] = update.message.text
    await update.message.reply_text("Descrição:", reply_markup=ReplyKeyboardRemove())
    return S_DESCRICAO

async def step_descricao(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Descrição"] = update.message.text
    await update.message.reply_text("Credor (fornecedor/órgão):")
    return S_CREDOR

async def step_credor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Credor"] = update.message.text
    await update.message.reply_text("Valor (ex: 1500,00):")
    return S_VALOR

async def step_valor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Valor (R$)"] = update.message.text.replace("R$","").strip()
    await update.message.reply_text(
        "Vencimento:\n"
        "Digite a data (DD/MM/AAAA) ou escolha um atalho:",
        reply_markup=kb(["Hoje","Em 7 dias","Em 15 dias","Em 30 dias"])
    )
    return S_VENCIMENTO

async def step_vencimento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text.strip()
    atalhos = {"Hoje": 0, "Em 7 dias": 7, "Em 15 dias": 15, "Em 30 dias": 30}
    if txt in atalhos:
        from datetime import timedelta
        d = date.today() + timedelta(days=atalhos[txt])
    else:
        d = parse_date(txt)
    if not d:
        await update.message.reply_text("❌ Data inválida. Use DD/MM/AAAA:")
        return S_VENCIMENTO
    context.user_data["Vencimento"] = d.strftime("%d/%m/%Y")
    await update.message.reply_text("É recorrente?", reply_markup=kb(["Sim","Não"]))
    return S_RECORRENTE

async def step_recorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Recorrente"] = update.message.text
    if update.message.text == "Sim":
        await update.message.reply_text("Frequência:", reply_markup=kb(FREQ_OPT))
        return S_FREQUENCIA
    context.user_data["Frequência"] = "Única"
    await update.message.reply_text("Observação (ou /pular):", reply_markup=ReplyKeyboardRemove())
    return S_OBS

async def step_frequencia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Frequência"] = update.message.text
    await update.message.reply_text("Observação (ou /pular):", reply_markup=ReplyKeyboardRemove())
    return S_OBS

async def step_obs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["Observação"] = "" if update.message.text == "/pular" else update.message.text
    d = context.user_data
    resumo = (
        f"✅ *Confirmar lançamento?*\n\n"
        f"🏢 {d.get('Empresa')} | 📂 {d.get('Categoria')}\n"
        f"📝 {d.get('Descrição')}\n"
        f"🏦 {d.get('Credor')}\n"
        f"💰 {fmt_brl(d.get('Valor (R$)',0))} | 📅 {d.get('Vencimento')}\n"
        f"🔁 {d.get('Recorrente')} ({d.get('Frequência')})"
    )
    await update.message.reply_text(resumo, reply_markup=kb(["✅ Confirmar","❌ Cancelar"]), parse_mode="Markdown")
    return S_CONFIRMA

async def step_confirma(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "Confirmar" not in update.message.text:
        await update.message.reply_text("❌ Cancelado.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END
    try:
        sh = get_sheet()
        ws = sh.worksheet("Contas")
        cp_id = next_id(ws)
        now   = datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
        data  = {**context.user_data, "ID": cp_id, "Status": "Pendente",
                 "Lançado por": "Bot", "Criado em": now}
        append_conta(ws, data)
        await update.message.reply_text(f"✅ *{cp_id}* lançado!",
            reply_markup=ReplyKeyboardRemove(), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {e}", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

async def cmd_cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Operação cancelada.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ── /cp (rápido) ───────────────────────────────────────────────────────────────
async def cmd_cp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    args = context.args
    if len(args) < 5:
        await update.message.reply_text(
            "❌ Formato: `/cp Credor Valor DD/MM/AAAA Empresa Categoria`\n"
            "Use _ no lugar de espaços.", parse_mode="Markdown")
        return
    credor   = args[0].replace("_"," ")
    valor    = args[1].replace(",",".")
    d        = parse_date(args[2])
    empresa  = args[3].replace("_"," ")
    categoria= args[4].replace("_"," ")
    if not d:
        await update.message.reply_text("❌ Data inválida. Use DD/MM/AAAA."); return
    try:
        sh = get_sheet()
        ws = sh.worksheet("Contas")
        cp_id = next_id(ws)
        now   = datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
        data  = {"ID": cp_id, "Empresa": empresa, "Categoria": categoria,
                 "Descrição": f"{categoria} {credor}", "Credor": credor,
                 "Valor (R$)": float(valor), "Vencimento": d.strftime("%d/%m/%Y"),
                 "Status": "Pendente", "Recorrente": "Não", "Frequência": "Única",
                 "Lançado por": "Comando", "Criado em": now}
        append_conta(ws, data)
        await update.message.reply_text(
            f"✅ *{cp_id}* lançado!\n{credor} | {fmt_brl(valor)} | {d.strftime('%d/%m/%Y')} | {empresa}",
            parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {e}")

# ── /pagar ─────────────────────────────────────────────────────────────────────
async def cmd_pagar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    args = context.args
    if not args:
        await update.message.reply_text("Uso: `/pagar CP0001` ou `/pagar CP0001 15/06/2026 1500`", parse_mode="Markdown"); return
    cp_id    = args[0].upper()
    data_pag = args[1] if len(args) > 1 else date.today().strftime("%d/%m/%Y")
    val_pago = args[2] if len(args) > 2 else ""
    try:
        sh = get_sheet()
        ws = sh.worksheet("Contas")
        row_idx, row_data = update_status(ws, cp_id, "Pago", data_pag, val_pago)
        if not row_idx:
            await update.message.reply_text(f"❌ ID *{cp_id}* não encontrado.", parse_mode="Markdown"); return

        msg = f"✅ *{cp_id}* pago em {data_pag}."

        # Recorrente — gera próxima
        if row_data and len(row_data) > 12 and row_data[11] == "Sim" and row_data[12] != "Única":
            prox_venc = next_vencimento(row_data[6], row_data[12])
            if prox_venc:
                novo_id = next_id(ws)
                now = datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
                nova = {
                    "ID": novo_id, "Empresa": row_data[1], "Categoria": row_data[2],
                    "Descrição": row_data[3], "Credor": row_data[4],
                    "Valor (R$)": row_data[5], "Vencimento": prox_venc,
                    "Status": "Pendente", "Recorrente": row_data[11],
                    "Frequência": row_data[12], "Observação": row_data[13] if len(row_data)>13 else "",
                    "Lançado por": "Auto-recorrente", "Criado em": now
                }
                append_conta(ws, nova)
                msg += f"\n🔁 Próxima parcela criada: *{novo_id}* — vence {prox_venc}"

        await update.message.reply_text(msg, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {e}")

# ── /cancelar ─────────────────────────────────────────────────────────────────
async def cmd_cancelar_conta(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Uso: /cancelar CP0001 Motivo aqui
    """
    if not is_authorized(update): return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text(
            "Uso: `/cancelar CP0001 Motivo aqui`\nEx: `/cancelar CP0003 Lançamento duplicado`",
            parse_mode="Markdown")
        return
    cp_id  = args[0].upper()
    motivo = " ".join(args[1:])
    try:
        sh   = get_sheet()
        ws   = sh.worksheet("Contas")
        rows = ws.get_all_values()
        row_idx = None
        for i, row in enumerate(rows[1:], start=2):
            if row[0] == cp_id:
                row_idx = i
                break
        if not row_idx:
            await update.message.reply_text(f"❌ ID *{cp_id}* não encontrado.", parse_mode="Markdown")
            return
        # Coluna Q=17, R=18
        ws.update(f"Q{row_idx}:R{row_idx}", [["Sim", motivo]])
        await update.message.reply_text(
            f"🚫 *{cp_id}* cancelado.\nMotivo: _{motivo}_",
            parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {e}")

# ── /listar ────────────────────────────────────────────────────────────────────
async def cmd_listar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    try:
        sh = get_sheet()
        ws = sh.worksheet("Contas")
        contas    = get_all_contas(ws)
        pendentes = [c for c in contas if c.get("Status") != "Pago"]
        if not pendentes:
            await update.message.reply_text("✅ Nenhuma conta pendente!"); return
        pendentes.sort(key=lambda c: parse_date(c.get("Vencimento","")) or date.max)
        lines = ["📋 *Contas Pendentes*\n"]
        for c in pendentes[:20]:
            d  = days_until(c.get("Vencimento",""))
            em = status_emoji(c.get("Status",""), d)
            lines.append(f"{em} `{c.get('ID','')}` {c.get('Credor','')} | {fmt_brl(c.get('Valor (R$)',0))} | {c.get('Vencimento','')} | {c.get('Empresa','')}")
        if len(pendentes) > 20:
            lines.append(f"\n_...e mais {len(pendentes)-20} contas._")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {e}")

# ── Resumos ────────────────────────────────────────────────────────────────────
async def _send_resumo(update, contas, titulo):
    await update.message.reply_text(build_resumo(contas, titulo), parse_mode="Markdown")

async def cmd_resumo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    ws = get_sheet().worksheet("Contas")
    await _send_resumo(update, [c for c in get_all_contas(ws) if c.get("Status") != "Pago"], "Resumo Geral")

async def cmd_resumo_dia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    ws   = get_sheet().worksheet("Contas")
    hoje = date.today().strftime("%d/%m/%Y")
    await _send_resumo(update, [c for c in get_all_contas(ws) if c.get("Vencimento") == hoje and c.get("Status") != "Pago"], f"Vencendo hoje — {hoje}")

async def cmd_resumo_semana(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    ws = get_sheet().worksheet("Contas")
    await _send_resumo(update, [c for c in get_all_contas(ws)
        if days_until(c.get("Vencimento","")) is not None and 0 <= days_until(c.get("Vencimento","")) <= 7
        and c.get("Status") != "Pago"], "Próximos 7 dias")

async def cmd_resumo_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    ws   = get_sheet().worksheet("Contas")
    hoje = date.today()
    await _send_resumo(update, [c for c in get_all_contas(ws)
        if (lambda d: d and d.year == hoje.year and d.month == hoje.month)(parse_date(c.get("Vencimento","")))
        and c.get("Status") != "Pago"], f"Mês atual — {hoje.strftime('%m/%Y')}")

# ── Import XLSX ────────────────────────────────────────────────────────────────
async def handle_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    doc = update.message.document
    if not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Envie um arquivo .xlsx."); return
    await update.message.reply_text("⏳ Processando...")
    try:
        f    = await context.bot.get_file(doc.file_id)
        path = f"/tmp/{doc.file_name}"
        await f.download_to_drive(path)
        wb   = openpyxl.load_workbook(path)
        ws_xl= wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws_xl.iter_rows(min_row=1,max_row=1))]
        sh   = get_sheet()
        ws   = sh.worksheet("Contas")
        now  = datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
        count= 0
        for row in ws_xl.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            r    = dict(zip(headers, row))
            cp_id= next_id(ws)
            venc = r.get("Vencimento") or r.get("vencimento") or ""
            if hasattr(venc, "strftime"): venc = venc.strftime("%d/%m/%Y")
            data = {"ID": cp_id,
                "Empresa":   r.get("Empresa") or "",
                "Categoria": r.get("Categoria") or "Outros",
                "Descrição": r.get("Descrição") or r.get("Descricao") or "",
                "Credor":    r.get("Credor") or "",
                "Valor (R$)":float(str(r.get("Valor") or r.get("Valor (R$)") or 0).replace(",",".")),
                "Vencimento":str(venc),
                "Status":    r.get("Status") or "Pendente",
                "Recorrente":r.get("Recorrente") or "Não",
                "Frequência":r.get("Frequência") or r.get("Frequencia") or "Única",
                "Observação":r.get("Observação") or "",
                "Lançado por":"Upload", "Criado em": now}
            append_conta(ws, data)
            count += 1
        await update.message.reply_text(f"✅ {count} conta(s) importada(s)!")
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {e}")

# ── /ajuda ─────────────────────────────────────────────────────────────────────
async def cmd_ajuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_authorized(update): return
    await update.message.reply_text(
        "*📚 NK Contas a Pagar — Comandos*\n\n"
        "*/nova* — guiado passo a passo\n"
        "`/cp Credor Valor Data Empresa Cat` — rápido\n\n"
        "*/pagar CP0001* — marca pago hoje\n"
        "`/pagar CP0001 15/06/2026 1500` — com data e valor\n\n"
        "*/listar* — pendentes por vencimento\n"
        "*/resumo* | */resumo\\_dia* | */resumo\\_semana* | */resumo\\_mes*\n\n"
        "📎 Envie `.xlsx` para importar em lote\n\n"
        "🔁 Contas recorrentes geram próxima parcela automaticamente ao pagar",
        parse_mode="Markdown"
    )

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    try:
        sh = get_sheet()
        ensure_sheets(sh)
        log.info("Google Sheets conectado.")
    except Exception as e:
        log.error("Erro Sheets: %s", e)

    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("nova", cmd_nova)],
        states={
            S_EMPRESA:    [MessageHandler(filters.TEXT & ~filters.COMMAND, step_empresa)],
            S_CATEGORIA:  [MessageHandler(filters.TEXT & ~filters.COMMAND, step_categoria)],
            S_DESCRICAO:  [MessageHandler(filters.TEXT & ~filters.COMMAND, step_descricao)],
            S_CREDOR:     [MessageHandler(filters.TEXT & ~filters.COMMAND, step_credor)],
            S_VALOR:      [MessageHandler(filters.TEXT & ~filters.COMMAND, step_valor)],
            S_VENCIMENTO: [MessageHandler(filters.TEXT & ~filters.COMMAND, step_vencimento)],
            S_RECORRENTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, step_recorrente)],
            S_FREQUENCIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, step_frequencia)],
            S_OBS:        [MessageHandler(filters.TEXT, step_obs)],
            S_CONFIRMA:   [MessageHandler(filters.TEXT & ~filters.COMMAND, step_confirma)],
        },
        fallbacks=[CommandHandler("cancelar", cmd_cancelar)],
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("start",         cmd_start))
    app.add_handler(CommandHandler("ajuda",         cmd_ajuda))
    app.add_handler(CommandHandler("cp",            cmd_cp))
    app.add_handler(CommandHandler("pagar",         cmd_pagar))
    app.add_handler(CommandHandler("cancelar",      cmd_cancelar_conta))
    app.add_handler(CommandHandler("listar",        cmd_listar))
    app.add_handler(CommandHandler("resumo",        cmd_resumo))
    app.add_handler(CommandHandler("resumo_dia",    cmd_resumo_dia))
    app.add_handler(CommandHandler("resumo_semana", cmd_resumo_semana))
    app.add_handler(CommandHandler("resumo_mes",    cmd_resumo_mes))
    app.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), handle_xlsx))

    jq = app.job_queue
    jq.run_repeating(job_alertas, interval=3600, first=10)
    jq.run_daily(job_resumo_diario,
        time=datetime.strptime("07:00","%H:%M").replace(tzinfo=TZ).timetz())

    log.info("Bot iniciado.")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
